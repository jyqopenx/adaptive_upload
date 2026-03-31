from __future__ import annotations

from copy import copy
from io import BytesIO
import zipfile

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def copy_sheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)

    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)


def write_df_to_sheet(df: pd.DataFrame, ws):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return output.getvalue()


def workbook_to_bytes(wb: Workbook) -> bytes:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_output_workbook(source_instruction_ws, data_sheet_name: str, df: pd.DataFrame) -> bytes:
    out_wb = Workbook()

    ws1 = out_wb.active
    ws1.title = source_instruction_ws.title
    copy_sheet(source_instruction_ws, ws1)

    ws2 = out_wb.create_sheet(title=data_sheet_name)
    write_df_to_sheet(df, ws2)

    return workbook_to_bytes(out_wb)


def process_revenue_supply_files(
    instructions_file,
    prior_pubid_file,
    supply_data_file,
    selected_month_start,
):
    instructions_wb = load_workbook(instructions_file)

    prior_wb = load_workbook(prior_pubid_file)

    prior_pubid = pd.read_excel(prior_pubid_file, sheet_name=0, header=0)
    mrr_mapping = pd.read_excel(prior_pubid_file, sheet_name=1, header=0)

    supply = pd.read_csv(supply_data_file, encoding="latin1")

    conditions = [
        supply["AD_FORMAT"] == "(null)",
        supply["AD_FORMAT"] == "BANNER",
    ]
    choices = [
        "(null)",
        "Display",
    ]
    supply["new_format"] = np.select(
        conditions,
        choices,
        default=supply["AD_FORMAT"].astype(str) + " - " + supply["VIDEO_FORMAT"].astype(str),
    )

    supply.columns = supply.columns.str.strip().str.lower()

    text_cols = [
        "device_type",
        "environment",
        "partner_1",
        "level",
        "sf_account_id",
        "integration",
        "ad_format",
        "transaction_type",
        "bidout_partner",
        "feature",
    ]

    for col in text_cols:
        if col in supply.columns:
            supply[col] = supply[col].astype(str).str.strip()

    supply["device_type"] = supply["device_type"].str.lower()
    supply["environment"] = supply["environment"].str.lower()

    supply["partner_1"] = supply["partner_1"].replace(["", "nan", "none"], pd.NA)
    supply["partner_1"] = supply["partner_1"].fillna("(blank)")

    agg_cols = {
        "exchange_requests": "sum",
        "tot_mkt_impressions": "sum",
        "tot_spend_usd": "sum",
        "exchange_net_revenue": "sum",
        "tot_partner_fee": "sum",
        "tot_exchange_net_revenue_final": "sum",
        "ms_net_revenue": "sum",
        "ssp_net_revenue": "sum",
    }

    for col in agg_cols:
        if col in supply.columns:
            supply[col] = pd.to_numeric(supply[col], errors="coerce")

    selected_month_start = pd.Timestamp(selected_month_start).replace(day=1)
    last_month_col_name = (
        f"{selected_month_start.month}/{selected_month_start.day}/{selected_month_start.year}"
    )
    month_label = selected_month_start.strftime("%Y%m")

    group_cols = [
        "level",
        "sf_account_id",
        "integration",
        "new_format",
        "transaction_type",
        "bidout_partner",
        "feature",
        "partner_1",
    ]

    def build_final_df(df: pd.DataFrame, device_type: str, environment: str) -> pd.DataFrame:
        filtered = df[
            (df["device_type"] == device_type.lower())
            & (df["environment"] == environment.lower())
        ].copy()

        grouped = filtered.groupby(group_cols, dropna=False).agg(agg_cols).reset_index()

        long_df = grouped.melt(
            id_vars=group_cols,
            value_vars=list(agg_cols.keys()),
            var_name="metric_name",
            value_name=last_month_col_name,
        )

        long_df["Account"] = "Sum of " + long_df["metric_name"]
        long_df["Level Code"] = "OpenX Tech"
        long_df["Publisher ID Code"] = long_df["sf_account_id"]
        long_df["Integration Code"] = long_df["integration"]
        long_df["Ad_Format Code"] = long_df["new_format"]
        long_df["Transaction_Type Code"] = long_df["transaction_type"]
        long_df["Bidout_Partner Code"] = long_df["bidout_partner"]
        long_df["Features Code"] = long_df["feature"]
        long_df["Partner 1 Code"] = long_df["partner_1"]

        result = long_df[
            [
                "Account",
                "Level Code",
                "Publisher ID Code",
                "Integration Code",
                "Ad_Format Code",
                "Transaction_Type Code",
                "Bidout_Partner Code",
                "Features Code",
                "Partner 1 Code",
                last_month_col_name,
            ]
        ].copy()

        result["Publisher ID Code"] = result["Publisher ID Code"].replace("nan", "(blank)")
        return result

    def create_ci_df(df: pd.DataFrame, month_col: str) -> pd.DataFrame:
        ci_df = df.iloc[:, 2:9].reset_index(drop=True)
        ci_df.insert(0, "Account", "trigger")
        ci_df.insert(1, "Level Code", "OpenX Tech")
        ci_df[month_col] = 1
        return ci_df

    desktop_web_final_df = build_final_df(supply, "desktop", "web")
    mobile_web_final_df = build_final_df(supply, "mobile", "web")
    mobile_app_final_df = build_final_df(supply, "mobile", "app")

    desktop_web_ci_df = create_ci_df(desktop_web_final_df, last_month_col_name)
    mobile_web_ci_df = create_ci_df(mobile_web_final_df, last_month_col_name)
    mobile_app_ci_df = create_ci_df(mobile_app_final_df, last_month_col_name)

    file_specs = [
        {
            "df": desktop_web_final_df,
            "instruction_sheet_index": 0,
            "file_name": f"_A_01_Rev_-_Core_-_Model_-_Desk LOAD FILE ({month_label}).xlsx",
            "data_sheet_name": "_A.01 Rev - Core - Model - Desk",
            "folder": "supply",
        },
        {
            "df": desktop_web_ci_df,
            "instruction_sheet_index": 0,
            "file_name": f"_A_01_Rev_-_Core_-_Model_-_Desk LOAD FILE ({month_label})- trigger.xlsx",
            "data_sheet_name": "_A.01 Rev - Core - Model - Desk",
            "folder": "supply",
        },
        {
            "df": mobile_web_final_df,
            "instruction_sheet_index": 1,
            "file_name": f"_A_02_Rev_-_Core_-_Model_-_Mobi LOAD FILE ({month_label}).xlsx",
            "data_sheet_name": "_A.02 Rev - Core - Model - Mobi",
            "folder": "supply",
        },
        {
            "df": mobile_web_ci_df,
            "instruction_sheet_index": 1,
            "file_name": f"_A_02_Rev_-_Core_-_Model_-_Mobi LOAD FILE ({month_label})- trigger.xlsx",
            "data_sheet_name": "_A.02 Rev - Core - Model - Mobi",
            "folder": "supply",
        },
        {
            "df": mobile_app_final_df,
            "instruction_sheet_index": 2,
            "file_name": f"_A_03_Rev_-_Core_-_Model_-_Mobi LOAD FILE ({month_label}).xlsx",
            "data_sheet_name": "_A.03 Rev - Core - Model - Mobi",
            "folder": "supply",
        },
        {
            "df": mobile_app_ci_df,
            "instruction_sheet_index": 2,
            "file_name": f"_A_03_Rev_-_Core_-_Model_-_Mobi LOAD FILE ({month_label}) - trigger.xlsx",
            "data_sheet_name": "_A.03 Rev - Core - Model - Mobi",
            "folder": "supply",
        },
    ]

    generated_reports: dict[str, bytes] = {}
    zip_entries: list[tuple[str, bytes]] = []

    for spec in file_specs:
        source_instruction_ws = instructions_wb.worksheets[spec["instruction_sheet_index"]]
        report_bytes = build_output_workbook(
            source_instruction_ws=source_instruction_ws,
            data_sheet_name=spec["data_sheet_name"],
            df=spec["df"],
        )
        generated_reports[spec["file_name"]] = report_bytes
        zip_entries.append((f'{spec["folder"]}/{spec["file_name"]}', report_bytes))

    new_group_cols = [
        "level",
        "sf_account_id",
        "integration",
        "new_format",
        "transaction_type",
        "environment",
        "device_type",
        "partner_1",
    ]

    consolidated_final_df = supply[new_group_cols].drop_duplicates().reset_index(drop=True)
    consolidated_final_df.insert(0, "Account", "trigger")
    consolidated_final_df.insert(1, "Level Code", "OpenX Tech")
    consolidated_final_df[last_month_col_name] = 1

    consolidated_final_df = consolidated_final_df.rename(
        columns={
            "sf_account_id": "Publisher ID Code",
            "integration": "Integration Code",
            "new_format": "Ad_Format Code",
            "transaction_type": "Transaction_Type Code",
            "environment": "Environment Code",
            "device_type": "Device_Type Code",
            "partner_1": "Partner 1 Code",
        }
    )

    consolidated_final_df = consolidated_final_df[
        [
            "Account",
            "Level Code",
            "Publisher ID Code",
            "Integration Code",
            "Ad_Format Code",
            "Transaction_Type Code",
            "Environment Code",
            "Device_Type Code",
            "Partner 1 Code",
            last_month_col_name,
        ]
    ].copy()

    consolidated_final_df["Publisher ID Code"] = consolidated_final_df["Publisher ID Code"].replace(
        "nan", "(blank)"
    )

    assumptions_file_1 = (
        f"A_01_Core_-_Assumptions_Summary LOAD FILE ({month_label})- TRIGGER.xlsx"
    )
    assumptions_bytes_1 = build_output_workbook(
        source_instruction_ws=instructions_wb.worksheets[3],
        data_sheet_name="A.01 Core - Assumptions Summary",
        df=consolidated_final_df,
    )
    generated_reports[assumptions_file_1] = assumptions_bytes_1
    zip_entries.append((f"assumptions/{assumptions_file_1}", assumptions_bytes_1))

    new_consolidated_group_cols = [
        "level",
        "sf_account_id",
        "integration",
        "device_type",
        "environment",
        "new_format",
        "feature",
        "bidout_partner",
    ]

    consolidated_trigger_df = supply[new_consolidated_group_cols].drop_duplicates().reset_index(
        drop=True
    )
    consolidated_trigger_df.insert(0, "Account", "trigger")

    consolidated_trigger_df = consolidated_trigger_df.rename(
        columns={
            "level": "Level Code",
            "sf_account_id": "Publisher ID Code",
            "integration": "Integration Code",
            "device_type": "Device_Category Code",
            "environment": "Environment Code",
            "new_format": "Ad_Format Code",
            "feature": "Features Code",
            "bidout_partner": "Bidout_Partner Code",
        }
    )

    consolidated_trigger_df[last_month_col_name] = 1

    consolidated_trigger_df = consolidated_trigger_df[
        [
            "Account",
            "Level Code",
            "Publisher ID Code",
            "Integration Code",
            "Device_Category Code",
            "Environment Code",
            "Ad_Format Code",
            "Features Code",
            "Bidout_Partner Code",
            last_month_col_name,
        ]
    ].copy()

    consolidated_trigger_df["Publisher ID Code"] = consolidated_trigger_df[
        "Publisher ID Code"
    ].replace("nan", "(blank)")

    assumptions_file_2 = f"_A_07_Consolidated_-_Rev_-_Core LOAD FILE ({month_label}) - TRIGGERa.xlsx"
    assumptions_bytes_2 = build_output_workbook(
        source_instruction_ws=instructions_wb.worksheets[4],
        data_sheet_name="_A.07 Consolidated - Rev - Core",
        df=consolidated_trigger_df,
    )
    generated_reports[assumptions_file_2] = assumptions_bytes_2
    zip_entries.append((f"assumptions/{assumptions_file_2}", assumptions_bytes_2))

    unique_publishers_supply = supply[
        [
            "sf_account_id",
            "sf_account_name",
            "publisher_type__c",
            "management_reporting_region__c",
        ]
    ].drop_duplicates().copy()

    prior_pubid_list = prior_pubid["pub_id"].astype(str).tolist()
    unique_publishers_supply["sf_account_id"] = unique_publishers_supply["sf_account_id"].astype(
        str
    )

    new_pub_ids_df = unique_publishers_supply[
        ~unique_publishers_supply["sf_account_id"].isin(prior_pubid_list)
    ].copy()

    mrr_mapping_merged = mrr_mapping.rename(columns={"MRR": "management_reporting_region__c"})

    new_pub_ids_df = pd.merge(
        new_pub_ids_df,
        mrr_mapping_merged[["management_reporting_region__c", "MRR Group"]],
        on="management_reporting_region__c",
        how="left",
    )

    new_pub_ids_df["Relationship Code"] = "O&O"

    final_new_pub_ids_df = new_pub_ids_df.rename(
        columns={
            "sf_account_id": "Dimension Value Name",
            "sf_account_name": "Publisher Name",
            "publisher_type__c": "Publisher Type",
            "management_reporting_region__c": "MRR Code",
            "MRR Group": "MRR_Group Code",
        }
    )

    final_new_pub_ids_df = final_new_pub_ids_df[
        [
            "Dimension Value Name",
            "Publisher Name",
            "Publisher Type",
            "MRR Code",
            "MRR_Group Code",
            "Relationship Code",
        ]
    ].copy()

    final_new_pub_ids_df["Monthly Total Publisher Cohort Name"] = selected_month_start.strftime(
        "%B %Y"
    )
    final_new_pub_ids_df["Annual Total Publisher Cohort Code"] = selected_month_start.year
    final_new_pub_ids_df[f"{selected_month_start.year} Cohort Code"] = "New"
    final_new_pub_ids_df["Publisher Status Code"] = "Active"
    final_new_pub_ids_df["Launch Cohort Code"] = final_new_pub_ids_df[
        "Annual Total Publisher Cohort Code"
    ]

    final_new_pub_ids_df = pd.merge(
        final_new_pub_ids_df,
        mrr_mapping_merged[["management_reporting_region__c", "Publisher Region"]],
        left_on="MRR Code",
        right_on="management_reporting_region__c",
        how="left",
    )

    final_new_pub_ids_df = final_new_pub_ids_df.rename(
        columns={"Publisher Region": "Publisher Region Code"}
    ).drop(columns=["management_reporting_region__c"])

    final_new_pub_ids_df = final_new_pub_ids_df.loc[
        :, ~final_new_pub_ids_df.columns.duplicated()
    ]

    dimensions_filename = "dimensions.xlsx"
    dimensions_bytes = dataframe_to_excel_bytes(final_new_pub_ids_df)
    generated_reports[dimensions_filename] = dimensions_bytes
    zip_entries.append(
        (f"Dimensions & Attributes Management/{dimensions_filename}", dimensions_bytes)
    )

    new_pub_ids_to_append = final_new_pub_ids_df[["Dimension Value Name"]].rename(
        columns={"Dimension Value Name": "pub_id"}
    )
    updated_prior_pubid_df = pd.concat([prior_pubid, new_pub_ids_to_append], ignore_index=True)

    updated_prior_pubid_wb = Workbook()
    ws_prior_1 = updated_prior_pubid_wb.active
    ws_prior_1.title = prior_wb.sheetnames[0]
    write_df_to_sheet(updated_prior_pubid_df, ws_prior_1)

    original_mrr_sheet = prior_wb.worksheets[1]
    ws_prior_2 = updated_prior_pubid_wb.create_sheet(title=original_mrr_sheet.title)
    copy_sheet(original_mrr_sheet, ws_prior_2)

    current_pubids_filename = "current_pubids.xlsx"
    current_pubids_bytes = workbook_to_bytes(updated_prior_pubid_wb)
    generated_reports[current_pubids_filename] = current_pubids_bytes
    zip_entries.append(
        (
            f"Dimensions & Attributes Management/{current_pubids_filename}",
            current_pubids_bytes,
        )
    )

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for zip_path, file_bytes in zip_entries:
            zf.writestr(zip_path, file_bytes)
    zip_buffer.seek(0)

    return {
        "generated_reports": generated_reports,
        "zip_bytes": zip_buffer.getvalue(),
        "month_label": month_label,
        "last_month_col_name": last_month_col_name,
        "new_publishers_df": final_new_pub_ids_df,
        "updated_prior_pubid_df": updated_prior_pubid_df,
        "desktop_web_final_df": desktop_web_final_df,
        "mobile_web_final_df": mobile_web_final_df,
        "mobile_app_final_df": mobile_app_final_df,
        "desktop_web_ci_df": desktop_web_ci_df,
        "mobile_web_ci_df": mobile_web_ci_df,
        "mobile_app_ci_df": mobile_app_ci_df,
        "consolidated_final_df": consolidated_final_df,
        "consolidated_trigger_df": consolidated_trigger_df,
    }
